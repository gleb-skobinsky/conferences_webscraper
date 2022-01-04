from selenium import webdriver
from selenium.webdriver.edge.service import Service
from bs4 import BeautifulSoup
from selenium.webdriver.support.ui import WebDriverWait
import pandas as pd
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
import math
import os
from openpyxl import Workbook
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.remote.webelement import WebElement
import PIL
from io import BytesIO
import numpy as np
import selenium
def hover_map(element: WebElement):
    location = element.location
    size = element.size
    png = driver.get_screenshot_as_png()
    raw_screenshot = PIL.Image.open(BytesIO(png))
    raw_screenshot = raw_screenshot.resize((1536, 720), PIL.Image.LANCZOS)
    raw_screenshot.save('test_raw.png')
    left = location['x']
    top = location['y'] - 500
    right = location['x'] + size['width']
    bottom = location['y'] -500 + size['height']
    im = raw_screenshot.crop((left, top, right, bottom))
    im.save('test.png')
    to_locate = PIL.Image.open('test.png')
    r=0
    g=1
    b=2
    r_query = 170
    g_query = 41
    b_query = 143
    image_array = np.array(to_locate)
    numpy_coordinates = np.where((image_array[:,:,r]==r_query)&(image_array[:,:,g] == g_query) & (image_array[:,:,b] == b_query))
    coordinates = [numpy_coordinates[0], numpy_coordinates[1]]
    
    first_coord_y = int(coordinates[0][0])
    first_coord_x = int(coordinates[1][0])
    
    
    move_mouse = ActionChains(driver).move_to_element_with_offset(element, first_coord_x, first_coord_y)
    move_mouse.perform()
    content_address = driver.page_source
    soup_address = BeautifulSoup(content_address, features="html.parser")
    local_address_list = soup_address.findAll('div', attrs={'class':'infobox-title', 'data-tag':'SDK.Infobox.Title'})
    return(local_address_list)
start = time.time()
s = Service(r"edgedriver_win64\msedgedriver.exe")
driver = webdriver.Edge(service=s)
address_list = []
titles = []
dates = []
locations = []
fields_of_interests = []
descriptions = []
sponsors = []
driver.get("https://conferences.ieee.org/conferences_events/conferences/search?q=*")
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'refine-search-wrap')))
content = driver.page_source
soup = BeautifulSoup(content, features="html.parser")
stronglist = soup.findAll('strong', attrs={'_ngcontent-c2':''})
num_conf = int(str(stronglist[3])[25:-9])
num_pages = math.ceil(num_conf/10)
for i in range(num_pages):
    
    driver.get("https://conferences.ieee.org/conferences_events/conferences/search?q=*&pos=" + str(i))
    driver.maximize_window()
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'refine-search-wrap')))
    content = driver.page_source
    soup = BeautifulSoup(content, features="html.parser")
    links_on_page = driver.find_elements(By.TAG_NAME, 'h4')
    virtual_status = []
    for art in soup.findAll('article', attrs={'_ngcontent-c2':''}):
        virtual_status.append(str(art))
    for art in soup.findAll('article', attrs={'_ngcontent-c2':''}):
        virtual_status.append(str(art))
    for conf in soup.findAll('h4', attrs={'_ngcontent-c2':''}):
        titles.append(str(conf)[21:-5])
    for date in soup.findAll('div', attrs={'class':'item-date'}):
        date_start = int(str(date).find('date">\n\t\t\t\t\t\t\t')) + 14
        date_end = str(date).find('\n\t\t\t\t\t\t\t<span _ngcontent')
        date_edit = str(date)[date_start: date_end]
        dates.append(date_edit)
    for iteration in range(int(len(virtual_status)/2)):
        local_interest_list = []
        for interest in driver.find_elements(By.XPATH, "//div[@class='item-details item-about']//span[@_ngcontent-c2='']//span[@_ngcontent-c2='']"):
            if str(interest.text) in virtual_status[iteration] and not str(interest.text)==';':
                if str(interest.text) not in local_interest_list:
                    local_interest_list.append(str(interest.text))
        fields_of_interests.append(local_interest_list)
    for instance in range(len(links_on_page)):
        links_on_page = driver.find_elements(By.TAG_NAME, 'h4')
        if 'Virtual Conference' in virtual_status[instance]:
            address_list.append('Virtual Conference')
            link = links_on_page[instance]
            window_before = driver.window_handles[0]
            driver.execute_script("window.scrollTo(0, 2000)")
            link.click()
            driver.execute_script("window.open()")
            window_after = driver.window_handles[1]
            driver.switch_to.window(window_after)
            driver.execute_script("window.close()")
            driver.switch_to.window(window_before)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@class='conference-details']//span[@_ngcontent-c8='']")))
            time.sleep(2)
            descr_el = soup_cfpg.findAll('p', attrs={'_ngcontent-c8':''})
            descr_el = str(descr_el[8])
            descr_el = descr_el.replace('<p _ngcontent-c8="">', '')
            descr_el = descr_el.replace('</p>', '')
            descriptions.append(descr_el)
            sponsors_el = driver.find_elements(By.XPATH, "//div[@class='conference-details']//span[@_ngcontent-c8='']")
            sponsors_local = []
            for sp in sponsors_el:
                sponsors_local.append(str(sp.get_attribute('innerHTML')))
            sponsors_index = sponsors_local.index("Sponsors:") + 1
            sponsors.append(str(sponsors_el[sponsors_index].get_attribute('innerHTML')))
            driver.execute_script("window.history.go(-1)")
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'refine-search-wrap')))
        else:
            link = links_on_page[instance]
            window_before = driver.window_handles[0]
            driver.execute_script("window.scrollTo(0, 2000)")
            link.click()
            driver.execute_script("window.open()")
            window_after = driver.window_handles[1]
            driver.switch_to.window(window_after)
            driver.execute_script("window.close()")
            driver.switch_to.window(window_before)
            try:
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'bm_LogoContainer quadrantOverride')]")))
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "labelCanvasId")))
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'Microsoft.Maps.Imagery.RoadSceneWithoutLabels')))
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'mapFocus')))
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//canvas[contains(@style, 'z-index: -1; position: absolute; opacity: 0;')]")))
                driver.execute_script("window.scrollTo(0, 500)") 
                map = driver.find_element(By.ID, 'Microsoft.Maps.Imagery.RoadSceneWithoutLabels')
                move_mouse = ActionChains(driver).move_to_element_with_offset(map, 471, 250)
                move_mouse.perform()
                content_cfpg = driver.page_source
                soup_cfpg = BeautifulSoup(content_cfpg, features="html.parser")
                local_address_list = soup_cfpg.findAll('div', attrs={'class':'infobox-title', 'data-tag':'SDK.Infobox.Title'})
                if len(local_address_list) == 0:
                    local_address_list = hover_map(map)
                address_list.append(str(local_address_list[0])[56:-6])
                descr_el = soup_cfpg.findAll('p', attrs={'_ngcontent-c8':''})
                descr_el = str(descr_el[8])
                descr_el = descr_el.replace('<p _ngcontent-c8="">', '')
                descr_el = descr_el.replace('</p>', '')
                descriptions.append(descr_el)
                sponsors_el = driver.find_elements(By.XPATH, "//div[@class='conference-details']//span[@_ngcontent-c8='']")
                sponsors_local = []
                for sp in sponsors_el:
                    sponsors_local.append(str(sp.get_attribute('innerHTML')))
                sponsors_index = sponsors_local.index("Sponsors:") + 1
                sponsors.append(str(sponsors_el[sponsors_index].get_attribute('innerHTML')))
                driver.execute_script("window.history.go(-1)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'refine-search-wrap')))
            except selenium.common.exceptions.TimeoutException:
                address_list.append('Address unknown')
                descr_el = soup_cfpg.findAll('p', attrs={'_ngcontent-c8':''})
                descr_el = str(descr_el[8])
                descr_el = descr_el.replace('<p _ngcontent-c8="">', '')
                descr_el = descr_el.replace('</p>', '')
                descriptions.append(descr_el)
                sponsors_el = driver.find_elements(By.XPATH, "//div[@class='conference-details']//span[@_ngcontent-c8='']")
                sponsors_local = []
                for sp in sponsors_el:
                    sponsors_local.append(str(sp.get_attribute('innerHTML')))
                sponsors_index = sponsors_local.index("Sponsors:") + 1
                sponsors.append(str(sponsors_el[sponsors_index].get_attribute('innerHTML')))
                driver.execute_script("window.history.go(-1)")
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'refine-search-wrap')))
        print('New page # ' + str(i))
        print(len(titles))
        print(len(dates))
        print(len(fields_of_interests))
        print(len(address_list))
        print(len(sponsors))
        print(len(descriptions))
driver.quit()
fields = []
for list in fields_of_interests:
    string = ', '.join(list)
    fields.append(string)
print('Sponsors array length:', len(sponsors))
titles_dict = {'Conference titles': titles, 'Conference dates': dates, 'Fields of interests': fields, 'Address': address_list, 'Organizers': sponsors, 'Description': descriptions}
df = pd.DataFrame(titles_dict)
file = r'webscrape.xlsx'
while os.path.isfile(file):
    file = file[:-5] + ' (1)' + file[-5:]
else:
    wb = Workbook()
    wb.save(filename = file)
book = load_workbook(file)
writer = pd.ExcelWriter(file, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}
for sheetname in writer.sheets:
    df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= True)
writer.save()
end = time.time()
print(end - start)


